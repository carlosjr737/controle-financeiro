"""Importa a FATURA do banco (arquivo .xlsx/.csv exportado) direto pra aba 'Fatura [mês]'.

Isso resolve de vez o 'não bate': em vez de reconstruir a fatura pelas transações
(que nunca casa por causa de dólar/IOF/parcelas/postagem), usa o número REAL do banco.
- O cartão da aba vira EXATAMENTE a fatura (total do banco).
- Categorias vêm do seu histórico; o que não casa vai pra 'Outros'.
- Seus Pix/dinheiro (linhas manuais) são preservados.

SEGURANÇA: --backup duplica a aba antes; sem --apply é só SIMULAÇÃO.

Uso:
    python -m scripts.importar_fatura "/caminho/Fatura2026-07-15.xlsx" "Fatura Jun"
    python -m scripts.importar_fatura "/caminho/Fatura...xlsx" "Fatura Jun" --backup --apply
"""
import os
import sys
import csv
import datetime
import unicodedata
from collections import Counter, defaultdict


def _carregar_env(caminho=".env"):
    if not os.path.exists(caminho):
        return
    with open(caminho, encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if ln and not ln.startswith("#") and "=" in ln:
                k, v = ln.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


_carregar_env()
from deploy.sheets_adapter import _abrir_planilha, _parse_num   # noqa: E402

HEADER = ["Data", "Estabelecimento", "Portador", "Valor", "Parcela",
          "Classificação", "Status", "of_id"]


def _norm(s):
    s = (str(s) if s is not None else "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _ler_fatura(caminho):
    """Lê o arquivo da fatura. Colunas: Data; Estabelecimento; Portador; Valor; Parcela."""
    linhas = []
    if caminho.lower().endswith(".csv"):
        with open(caminho, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f, delimiter=";"):
                linhas.append((row.get("Data"), row.get("Estabelecimento"),
                               row.get("Portador"), _parse_num(row.get("Valor")),
                               row.get("Parcela")))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            r = list(r) + [None] * 5
            if r[0] is not None and r[3] is not None:
                d = r[0].strftime("%d/%m/%Y") if isinstance(r[0], datetime.datetime) else str(r[0])
                linhas.append((d, r[1], r[2], _parse_num(r[3]), r[4]))
    return [x for x in linhas if x[3] is not None]


def _mapa_classificacao(pl):
    """estabelecimento -> classificação, aprendido das abas Fatura existentes."""
    votos = defaultdict(Counter)
    for ws in pl.worksheets():
        if not ws.title.startswith("Fatura"):
            continue
        for r in ws.get_all_values()[1:]:
            if len(r) > 5 and r[1].strip() and r[5].strip():
                votos[_norm(r[1])][r[5].strip()] += 1
    exato = {k: v.most_common(1)[0][0] for k, v in votos.items()}
    pref = defaultdict(Counter)
    for k, v in votos.items():
        p = (k.split()[0][:8] if k.split() else k)
        pref[p][v.most_common(1)[0][0]] += sum(v.values())
    prefixo = {k: v.most_common(1)[0][0] for k, v in pref.items()}
    return exato, prefixo


def _classificar(estab, exato, prefixo):
    n = _norm(estab)
    c = exato.get(n)
    if not c:
        p = (n.split()[0][:8] if n.split() else n)
        c = prefixo.get(p)
    if not c or "estorn" in _norm(c):   # estorno depende de contexto -> não usa aqui
        return "Outros"
    return c


def _eh_manual_preservar(row):
    # mantém Pix/fixos (estab==classif) e dinheiro (começa com número)
    estab = row[1] if len(row) > 1 else ""
    classif = row[5] if len(row) > 5 else ""
    import re
    return (_norm(estab) == _norm(classif) and _norm(classif)) or bool(re.match(r"^\s*\d", str(estab or "")))


def main(caminho, aba, apply, fazer_backup):
    print(f"### Importar fatura -> '{aba}' — {'APLICANDO' if apply else 'SIMULAÇÃO'} ###")
    pl = _abrir_planilha()
    exato, prefixo = _mapa_classificacao(pl)
    fatura = _ler_fatura(caminho)
    total = sum(x[3] for x in fatura)
    print(f"Fatura lida: {len(fatura)} lançamentos | TOTAL R$ {total:,.2f}")

    # preserva os manuais Pix/dinheiro da aba atual
    ws = pl.worksheet(aba)
    manuais = []
    for r in ws.get_all_values()[1:]:
        r = list(r) + [""] * (8 - len(r))
        if any(str(c).strip() for c in r) and str(r[6]).strip().upper() != "OF" and _eh_manual_preservar(r):
            manuais.append(r)

    por_cat = Counter()
    novas_of = []
    for d, estab, port, val, parc in fatura:
        c = _classificar(estab, exato, prefixo)
        por_cat[c] += val
        ofid = f"fat-{d}-{_norm(estab)[:12]}-{val}"
        novas_of.append([d, estab, port or "", val, parc if parc not in (None, "") else "",
                         c, "OF", ofid])

    print(f"\nManuais preservados: {len(manuais)} | cartão (fatura): {len(novas_of)}")
    print("Cartão por categoria:")
    for c, v in por_cat.most_common():
        print(f"   {c[:26]:<26} R$ {v:>10,.2f}")

    if not apply:
        print("\n>> SIMULAÇÃO. Aplicar: acrescente --backup --apply")
        return

    if fazer_backup and f"BKP2 {aba}" not in {w.title for w in pl.worksheets()}:
        pl.duplicate_sheet(ws.id, new_sheet_name=f"BKP2 {aba}")
        print(f"backup: BKP2 {aba}")
    ws.clear()
    dados = [HEADER] + [m[:8] for m in manuais] + novas_of
    ws.append_rows(dados, value_input_option="USER_ENTERED")
    print(f"\n✅ '{aba}' = {len(manuais)} manuais + {len(novas_of)} cartão. "
          f"Cartão bate a fatura: R$ {total:,.2f}")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    caminho = args[0] if args else ""
    aba = args[1] if len(args) > 1 else "Fatura Jun"
    main(caminho, aba, apply="--apply" in sys.argv, fazer_backup="--backup" in sys.argv)
