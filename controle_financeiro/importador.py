import openpyxl
from controle_financeiro.models import Categoria, Transacao
from controle_financeiro.regras_negocio import eh_estorno

def _garantir_categoria(sessao, nome: str) -> Categoria:
    cat = sessao.query(Categoria).filter_by(nome=nome).one_or_none()
    if cat is None:
        cat = Categoria(nome=nome)
        sessao.add(cat); sessao.flush()
    return cat

def importar_historico(caminho_xlsx: str, sessao) -> None:
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)

    # cache {nome: Categoria} carregado de uma vez — evita SELECT por transação
    cache = {c.nome: c for c in sessao.query(Categoria).all()}

    def garantir(nome: str) -> Categoria:
        nome = nome.strip()
        cat = cache.get(nome)
        if cat is None:
            cat = Categoria(nome=nome)
            sessao.add(cat); sessao.flush()
            cache[nome] = cat
        return cat

    # 1) vocabulário de categorias
    if "Classificações" in wb.sheetnames:
        for (nome,) in wb["Classificações"].iter_rows(min_row=2, values_only=True):
            if nome and str(nome).strip():
                garantir(str(nome))

    # 2) transações das abas "Fatura ..."
    for nome_aba in wb.sheetnames:
        if not nome_aba.startswith("Fatura"):
            continue
        ws = wb[nome_aba]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] in (None, "", "Estabelecimento"):
                continue
            data, estab, portador, valor, parcela, classificacao, *_ = list(row) + [None] * 7
            if valor is None or estab is None:
                continue
            if eh_estorno(valor, classificacao):
                sessao.add(Transacao(estabelecimento=str(estab), portador=portador,
                                     valor=float(valor), parcela=_str(parcela),
                                     status_classificacao="estorno",
                                     mes_competencia=_competencia(data)))
                continue
            cat = garantir(str(classificacao)) if classificacao else None
            sessao.add(Transacao(
                estabelecimento=str(estab), portador=portador, valor=float(valor),
                parcela=_str(parcela), categoria_id=cat.id if cat else None,
                status_classificacao="confirmada" if cat else "pendente",
                confianca=1.0 if cat else 0.0, mes_competencia=_competencia(data)))
    sessao.commit()

def _str(v):
    return None if v is None else str(v)

def _competencia(data) -> str | None:
    if data is None:
        return None
    txt = str(data)
    return txt[:7] if len(txt) >= 7 else None
